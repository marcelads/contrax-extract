"""
Leitor_pdf.py
=============
Extrai dados de contracheques (folha mensal)
e gera uma planilha .xlsx estruturada.

Dependências:
    pip install pdfplumber openpyxl Pillow
"""
import re
import sys
import threading
import unicodedata
from datetime import datetime
from pathlib import Path

try:
    import pdfplumber
except ImportError:
    sys.exit("Instale pdfplumber:  pip install pdfplumber")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Instale openpyxl:  pip install openpyxl")

import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk

# ── Tema ───────────────────────────────────────────────────────────────────────
BG      = "#F5F2F5"
SURF    = "#B4B0B0"
SURF2   = "#C5C3C3"
BORDER  = "#DBDADA"
ACCENT  = "#700aa0"
ACCENT2 = "#770f8b"
TEXT    = "#1C1C1C"
MUTED   = "#888888"
SUCCESS = "#1a9e6e"
WARNING = "#0ca683"
ERROR   = "#B45BC0"
LOG_BG  = "#F8F8F8"
LOG_FG  = "#444444"

VALOR_RE = r"[\d\.]+,\d{2}"


def _norm(s: str) -> str:
    return unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode("ascii").lower()


def parse_valor(s):
    if not s:
        return None
    try:
        return float(s.replace(".", "").replace(",", "."))
    except ValueError:
        return None


def buscar_rubrica(texto: str, padroes: list):
    """Retorna o último valor monetário da primeira linha que bate qualquer padrão."""
    linhas_orig = texto.split("\n")
    linhas_norm = [_norm(l) for l in linhas_orig]
    for orig, norm in zip(linhas_orig, linhas_norm):
        for p in padroes:
            if re.search(p, norm):
                vals = re.findall(VALOR_RE, orig)
                if vals:
                    return parse_valor(vals[-1])
    return None


def extrair_contracheques(pdf_path: str, log_fn=print) -> list:
    vistos = set()
    funcionarios = []

    with pdfplumber.open(pdf_path) as pdf:
        total_pag = len(pdf.pages)
        for num_pag, page in enumerate(pdf.pages, 1):
            log_fn(f"  Lendo página {num_pag}/{total_pag}...")
            texto = page.extract_text()
            if not texto:
                continue

            m = re.search(
                r"C[oó]digo\s+Nome\s+do\s+Funcion[aá]rio\s+CBO\s+Departamento\s+Filial\s*\n"
                r"(\d+)\s+(.+?)\s+(\d{5,6})\s+\d+\s+\d+\s*\n"
                r"(.+?)\s+Admiss[aã]o:\s*(\d{2}/\d{2}/\d{4})",
                texto, re.IGNORECASE
            )
            if not m:
                m = re.search(
                    r"(\d{2,4})\s+([A-ZÁÉÍÓÚÀÂÊÔÃÕÜÇÑ][A-ZÁÉÍÓÚÀÂÊÔÃÕÜÇÑ\s]{5,}?)\s+(\d{5,6})\s+\d+\s+\d+\s*\n"
                    r"(.+?)\s+Admiss[aã]o:\s*(\d{2}/\d{2}/\d{4})",
                    texto, re.IGNORECASE
                )
            if not m:
                continue

            codigo   = int(m.group(1))
            nome     = m.group(2).strip()
            cbo      = m.group(3)
            cargo    = m.group(4).strip()
            admissao = m.group(5)

            if codigo in vistos:
                continue
            vistos.add(codigo)

            _tot_matches = re.findall(
                r"Sal[aá]rio\s+Base\s+Sal\.\s*Contr\.\s*INSS\s+Base\s+C[aá]lc\.\s*FGTS"
                r"\s+F\.?G\.?T\.?S\.?\s+do\s+M[eê]s\s+Base\s+C[aá]lc\.\s*IRRF\s+Faixa\s+IRRF\s*\n"
                r"(" + VALOR_RE + r")\s+"
                r"(" + VALOR_RE + r")\s+"
                r"(" + VALOR_RE + r")\s+"
                r"(" + VALOR_RE + r")\s+"
                r"(" + VALOR_RE + r")\s+"
                r"(" + VALOR_RE + r")",
                texto, re.IGNORECASE
            )
            _tot = _tot_matches[-1] if _tot_matches else None
            salario_base = parse_valor(_tot[0]) if _tot else None
            fgts         = parse_valor(_tot[3]) if _tot else None

            _liq_matches = re.findall(r"Valor\s+L[ií]quido\s+(" + VALOR_RE + r")", texto, re.IGNORECASE)
            valor_liquido = parse_valor(_liq_matches[-1]) if _liq_matches else None

            m_prov = re.search(r"Total\s+(?:de\s+)?(?:Proventos|Vencimentos)\s+(" + VALOR_RE + r")", texto, re.IGNORECASE)
            m_desc_tot = re.search(r"Total\s+(?:de\s+)?Descontos\s+(" + VALOR_RE + r")", texto, re.IGNORECASE)
            total_proventos = parse_valor(m_prov.group(1)) if m_prov else None
            total_descontos = parse_valor(m_desc_tot.group(1)) if m_desc_tot else None
            if total_proventos is None and valor_liquido is not None and total_descontos:
                total_proventos = round(valor_liquido + total_descontos, 2)

            gerencia = buscar_rubrica(texto, [
                r"ger[eê]ncia\b", r"adicional\s+de?\s*ger[eê]ncia", r"bonus\s*ger[eê]ncia",
                r"grat\.?\s*ger[eê]ncia", r"ger[eê]nte\b"
            ])
            supervisao = buscar_rubrica(texto, [
                r"supervis[aã]o\b", r"adicional\s+supervis", r"bonus\s*supervis",
                r"grat\.?\s*supervis", r"supervisor"
            ])
            horas_extras = buscar_rubrica(texto, [
                r"hora[s]?\s+extra[s]?\b", r"\bh\.?\s*extra[s]?\b", r"\bhe\s+\d",
                r"hora[s]?\s+exc[e]?d", r"50%", r"100%"
            ])
            adic_tempo = buscar_rubrica(texto, [
                r"adicional\s+de\s+tempo", r"adic\.?\s*tempo\b", r"adicional\s+tempo",
                r"adic\.?\s+t\.?\s+serv", r"adicional\s+por\s+tempo",
                r"anuenio\b", r"quinquenio\b", r"triênio\b", r"trienio\b"
            ])
            inss = buscar_rubrica(texto, [
                r"\binss\b", r"prev\.?\s*social\b", r"i\.n\.s\.s\.", r"previdencia\s+social"
            ])
            irrf = buscar_rubrica(texto, [
                r"\birrf\b", r"imposto\s+de\s+renda\s+retido", r"i\.r\.r\.f\.", r"irpf\b"
            ])
            faltas = buscar_rubrica(texto, [
                r"\bfalta[s]?\b", r"desc\.?\s*falta", r"desconto\s+falta",
                r"atraso[s]?\b", r"desc\.?\s*atraso"
            ])
            vale_transp = buscar_rubrica(texto, [
                r"vale\s+transp", r"\bv\.?\s*t\.?\b", r"\bvt\b(?!\w)"
            ])
            vale_alim = buscar_rubrica(texto, [
                r"vale\s+aliment", r"vale\s+refei[cç]", r"alimenta[cç][aã]o"
            ])
            adic_noturno = buscar_rubrica(texto, [
                r"adicional\s+noturno", r"adic\.?\s*noturno",
                r"hora[s]?\s+noturna[s]?", r"\bnoturno\b", r"reflexo\s+adic",
            ])
            sal_familia = buscar_rubrica(texto, [
                r"sal[aá]rio\s+fam[ií]lia", r"sal\.?\s*fam[ií]lia", r"\bfam[ií]lia\b",
            ])
            sal_maternidade = buscar_rubrica(texto, [
                r"sal[aá]rio\s+maternidade", r"licen[cç]a\s+maternidade", r"maternidade\b",
            ])
            plano_saude = buscar_rubrica(texto, [
                r"plano\s+de\s+sa[uú]de", r"desc\.?\s*plano\s+sa[uú]de",
                r"assist[eê]ncia\s+m[eé]dica", r"8111\b",
            ])
            plano_odonto = buscar_rubrica(texto, [
                r"plano\s+odontol[oó]gico", r"odontol[oó]gico\b",
                r"odonto\b", r"desc\.?\s*plano\s+odonto",
            ])
            consignado = buscar_rubrica(texto, [
                r"emp\.?\s*cr[eé]d", r"cred\.?\s*trab", r"consignado\b",
                r"empr[eé]stimo\b", r"parcela\s+empr[eé]stimo", r"9750\b",
            ])
            adiant_decimo = buscar_rubrica(texto, [
                r"adiant\.?\s*13", r"13[oº]\s+sal[aá]rio",
                r"decimo\s+terceiro", r"d[eé]cimo\s+terceiro", r"1[ao]\s+parcela\s+13",
            ])
            adiantamento = buscar_rubrica(texto, [
                r"adiantamento\b", r"\badiant\b", r"1[ao]\s+parcela\b", r"antecipa[cç][aã]o"
            ])

            _d = dict(
                cod=codigo, nome=nome, cbo=cbo, cargo=cargo, admissao=admissao,
                salario_base=salario_base, gerencia=gerencia, supervisao=supervisao,
                horas_extras=horas_extras, adic_tempo=adic_tempo,
                adic_noturno=adic_noturno, sal_familia=sal_familia,
                sal_maternidade=sal_maternidade, total_proventos=total_proventos,
                inss=inss, irrf=irrf, faltas=faltas, vale_transp=vale_transp,
                vale_alim=vale_alim, plano_saude=plano_saude,
                plano_odonto=plano_odonto, consignado=consignado,
                adiant_decimo=adiant_decimo, adiantamento=adiantamento,
                total_descontos=total_descontos, liquido=valor_liquido, fgts=fgts,
            )
            for _k, _v in _d.items():
                if isinstance(_v, float):
                    _d[_k] = round(_v, 2)
            funcionarios.append(_d)

    funcionarios.sort(key=lambda x: x["cod"])
    return funcionarios


_KEYS = [
    "cod", "nome", "cbo", "cargo", "admissao",
    "salario_base", "gerencia", "supervisao", "horas_extras", "adic_tempo",
    "adic_noturno", "sal_familia", "sal_maternidade",
    "total_proventos",
    "inss", "irrf", "faltas", "vale_transp", "vale_alim",
    "plano_saude", "plano_odonto", "consignado", "adiant_decimo", "adiantamento",
    "total_descontos",
    "liquido", "fgts",
]
_HEADERS = [
    "Cód", "Nome", "CBO", "Cargo", "Admissão",
    "Salário Base", "Gerência", "Supervisão", "H. Extras", "Adic. Tempo Serv.",
    "Adic. Noturno", "Sal. Família", "Sal. Maternidade",
    "Total Proventos",
    "INSS", "IRRF", "Faltas", "Vale Transp.", "Vale Alim.",
    "Plano Saúde", "Plano Odonto", "Consignado", "Adiant. 13º", "Adiantamento",
    "Total Descontos",
    "Valor Líquido", "FGTS do Mês",
]
_WIDTHS = [
    6, 40, 9, 28, 12,
    14, 12, 12, 12, 17,
    14, 13, 17,
    16,
    12, 12, 12, 12, 12,
    13, 13, 13, 13, 14,
    16,
    14, 13,
]
_MONEY_COLS = set(range(6, 27))
_CENTER_COLS = {1, 3, 5}

_GRUPOS = [
    ("IDENTIFICAÇÃO", 5,  "1F4E79"),
    ("PROVENTOS",     9,  "1E6B3C"),
    ("DESCONTOS",     11, "7B2020"),
    ("TOTAIS",        2,  "2E4E7A"),
]


def validar_funcionarios(funcionarios: list, log_fn=print) -> list:
    alertas = []
    for f in funcionarios:
        cod  = f["cod"]
        nome = f["nome"]
        erros = []

        for campo in ("salario_base", "total_proventos", "total_descontos", "liquido", "fgts", "inss"):
            if f.get(campo) is None:
                erros.append(f"campo '{campo}' não extraído (None)")

        tp = f.get("total_proventos")
        td = f.get("total_descontos")
        lq = f.get("liquido")
        if tp is not None and td is not None and lq is not None:
            diff = abs(round(tp - td, 2) - round(lq, 2))
            if diff > 0.02:
                erros.append(
                    f"proventos({tp}) - descontos({td}) = {round(tp-td,2)} "
                    f"!= liquido({lq})  |diff={diff:.2f}|"
                )

        fg = f.get("fgts")
        if fg is not None and lq is not None and lq > 0 and fg > lq:
            erros.append(f"fgts({fg}) > liquido({lq}) — possível inversão")

        sb = f.get("salario_base")
        if sb is not None and sb <= 0:
            erros.append(f"salario_base={sb} inválido")

        if tp is not None and td is not None and td > tp:
            if lq != 0:
                erros.append(f"total_descontos({td}) > total_proventos({tp})")

        for campo, val in f.items():
            if isinstance(val, float):
                if round(val, 2) != round(val, 10):
                    erros.append(f"float sujo em '{campo}': {val}")
                f[campo] = round(val, 2)

        if erros:
            alertas.append({"cod": cod, "nome": nome, "erros": erros})
            for e in erros:
                log_fn(f"  [AVISO] [{cod}] {nome}: {e}")

    if not alertas:
        log_fn("  Validação OK — nenhuma inconsistência encontrada.")

    return alertas


def gerar_xlsx(funcionarios: list, output_path: str, alertas: list = None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Folha de Pagamento"

    thin  = Side(style="thin", color="CCCCCC")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill  = PatternFill("solid", start_color="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alt_fill  = PatternFill("solid", start_color="EBF3FB")
    branco    = PatternFill("solid", start_color="FFFFFF")
    tot_fill  = PatternFill("solid", start_color="D6E4F0")
    tot_font  = Font(name="Arial", bold=True, size=10)
    money_fmt = '#,##0.00'

    col_cur = 1
    for grp_nome, grp_cols, grp_cor in _GRUPOS:
        ws.merge_cells(start_row=1, start_column=col_cur,
                       end_row=1, end_column=col_cur + grp_cols - 1)
        c = ws.cell(row=1, column=col_cur, value=grp_nome)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", start_color=grp_cor)
        c.alignment = Alignment(horizontal="center", vertical="center")
        col_cur += grp_cols
    ws.row_dimensions[1].height = 20

    for i, (h, w) in enumerate(zip(_HEADERS, _WIDTHS), 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = hdr_align; c.border = borda
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 32

    for idx, f in enumerate(funcionarios, 3):
        fill = alt_fill if idx % 2 == 1 else branco
        for col, key in enumerate(_KEYS, 1):
            val = f.get(key)
            c = ws.cell(row=idx, column=col, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = fill; c.border = borda
            if col in _CENTER_COLS:
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col in _MONEY_COLS:
                c.alignment = Alignment(horizontal="right", vertical="center")
                if val is not None:
                    c.number_format = money_fmt
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")

    last_data = len(funcionarios) + 2
    tot_row   = last_data + 1

    lbl = ws.cell(row=tot_row, column=4, value="TOTAIS")
    lbl.font = Font(name="Arial", bold=True, color="1F4E79", size=10)
    lbl.alignment = Alignment(horizontal="right")

    cnt = ws.cell(row=tot_row, column=1, value=f"=COUNTA(A3:A{last_data})")
    cnt.font = tot_font; cnt.fill = tot_fill
    cnt.border = borda
    cnt.alignment = Alignment(horizontal="center")

    for col in _MONEY_COLS:
        ltr = get_column_letter(col)
        c = ws.cell(row=tot_row, column=col,
                    value=f"=SUM({ltr}3:{ltr}{last_data})")
        c.font = tot_font; c.fill = tot_fill
        c.border = borda; c.number_format = money_fmt
        c.alignment = Alignment(horizontal="right", vertical="center")

    ws.freeze_panes = "A3"

    if alertas:
        wa = wb.create_sheet("Alertas")
        wa.column_dimensions["A"].width = 8
        wa.column_dimensions["B"].width = 35
        wa.column_dimensions["C"].width = 80
        hdr_alerta = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        fill_alerta = PatternFill("solid", start_color="C00000")
        for col, txt in enumerate(["Cód", "Nome", "Problema"], 1):
            c = wa.cell(row=1, column=col, value=txt)
            c.font = hdr_alerta
            c.fill = fill_alerta
            c.alignment = Alignment(horizontal="center")
        row_a = 2
        for alerta in alertas:
            for erro in alerta["erros"]:
                wa.cell(row=row_a, column=1, value=alerta["cod"])
                wa.cell(row=row_a, column=2, value=alerta["nome"])
                wa.cell(row=row_a, column=3, value=erro)
                row_a += 1

    wb.save(output_path)


def gerar_nome_saida(pasta: Path, nome_base: str) -> Path:
    data_str = datetime.now().strftime("%d%m%Y")
    candidato = pasta / f"{nome_base}_{data_str}.xlsx"
    if not candidato.exists():
        return candidato
    n = 2
    while True:
        candidato = pasta / f"{nome_base}_{data_str}_v{n:02d}.xlsx"
        if not candidato.exists():
            return candidato
        n += 1


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Leitor de Contracheques")
        self.root.geometry("700x600")
        self.root.resizable(False, False)
        self.root.configure(bg=BG)
        self.pdf_var   = tk.StringVar()
        self.pasta_var = tk.StringVar()
        self._build_ui()

    def _build_ui(self):
        hdr = tk.Frame(self.root, bg=SURF, pady=12)
        hdr.pack(fill="x")

        txt_frame = tk.Frame(hdr, bg=SURF)
        txt_frame.pack(expand=True)
        tk.Label(txt_frame, text="LEITOR DE CONTRACHEQUES",
                 font=("Segoe UI", 13, "bold"),
                 bg=SURF, fg=ACCENT2).pack()
        tk.Label(txt_frame, text="Extrai dados da folha e gera planilha",
                 font=("Segoe UI", 9), bg=SURF, fg=MUTED).pack(pady=(3, 0))

        tk.Frame(self.root, bg=ACCENT, height=2).pack(fill="x")

        fp = tk.Frame(self.root, bg=BG, padx=26, pady=18)
        fp.pack(fill="x")

        self._campo(fp, "Arquivo PDF da folha de pagamento",
                    "Selecione o PDF com os contracheques",
                    self.pdf_var, self._browse_pdf)

        tk.Frame(fp, height=12, bg=BG).pack()

        self._campo(fp, "Pasta de destino da planilha",
                    "Onde salvar o arquivo .xlsx gerado",
                    self.pasta_var, self._browse_pasta)

        fb = tk.Frame(self.root, bg=BG, padx=26, pady=6)
        fb.pack(fill="x")
        self.btn_rodar = tk.Button(
            fb,
            text="▶   EXTRAIR DADOS",
            font=("Segoe UI", 11, "bold"),
            bg=ACCENT, fg="#FFFFFF",
            activebackground="#fc7c05",
            activeforeground="#FFFFFF",
            disabledforeground="#FFFFFF",
            relief="flat", cursor="hand2",
            command=self._iniciar
        )
        self.btn_rodar.pack(fill="x", ipady=12)

        tk.Frame(self.root, bg=BORDER, height=1).pack(fill="x", side="bottom")
        sb = tk.Frame(self.root, bg=SURF, padx=16, pady=6)
        sb.pack(fill="x", side="bottom")
        self.lbl_status = tk.Label(
            sb, text="Pronto",
            font=("Segoe UI", 9), bg=SURF, fg=MUTED, anchor="w"
        )
        self.lbl_status.pack(fill="x")

        fl = tk.Frame(self.root, bg=BG, padx=26, pady=0)
        fl.pack(fill="both", expand=True)
        tk.Label(fl, text="Log de execução",
                 font=("Segoe UI", 9, "bold"),
                 bg=BG, fg=MUTED).pack(anchor="w", pady=(12, 4))
        log_frame = tk.Frame(fl, bg=BORDER, padx=1, pady=1)
        log_frame.pack(fill="both", expand=True)
        self.log_text = tk.Text(
            log_frame,
            font=("Consolas", 9),
            bg=LOG_BG, fg=LOG_FG,
            insertbackground=TEXT,
            relief="flat", bd=0,
            state="disabled"
        )
        self.log_text.pack(fill="both", expand=True, ipady=8, ipadx=10)

    def _campo(self, parent, label, sublabel, var, cmd):
        tk.Label(parent, text=label,
                 font=("Segoe UI", 9, "bold"), bg=BG, fg=TEXT).pack(anchor="w")
        tk.Label(parent, text=sublabel,
                 font=("Segoe UI", 8), bg=BG, fg=MUTED).pack(anchor="w", pady=(2, 6))
        row = tk.Frame(parent, bg=BG)
        row.pack(fill="x")
        ef = tk.Frame(row, bg=BORDER, padx=1, pady=1)
        ef.pack(side="left", fill="x", expand=True)
        tk.Entry(ef, textvariable=var,
                 font=("Segoe UI", 9),
                 bg=SURF2, fg=TEXT,
                 insertbackground=TEXT,
                 relief="flat", bd=0
                 ).pack(fill="x", ipady=8, ipadx=10)
        tk.Frame(row, width=8, bg=BG).pack(side="left")
        tk.Button(row, text="Procurar",
                  font=("Segoe UI", 9, "bold"),
                  bg=SURF2, fg=ACCENT,
                  activebackground=BORDER, activeforeground=ACCENT,
                  relief="flat", cursor="hand2",
                  command=cmd
                  ).pack(side="left", ipady=9, ipadx=16)

    def _browse_pdf(self):
        f = filedialog.askopenfilename(
            title="Selecione o PDF da folha",
            filetypes=[("PDF", "*.pdf"), ("Todos", "*.*")]
        )
        if f:
            self.pdf_var.set(f)

    def _browse_pasta(self):
        d = filedialog.askdirectory(title="Selecione a pasta de destino")
        if d:
            self.pasta_var.set(d)

    def _set_status(self, msg, cor=None):
        self.lbl_status.config(text=msg, fg=cor or MUTED)
        self.root.update()

    def _log(self, msg=""):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", str(msg) + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")
        self.root.update()

    def _limpar_log(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

    def _iniciar(self):
        pdf_path = self.pdf_var.get().strip().strip('"')
        pasta    = self.pasta_var.get().strip().strip('"')

        if not pdf_path:
            messagebox.showerror("Erro", "Informe o caminho do PDF.")
            return
        if not Path(pdf_path).exists():
            messagebox.showerror("Erro", f"Arquivo não encontrado:\n{pdf_path}")
            return
        if not pasta:
            pasta = str(Path(pdf_path).parent)
            self.pasta_var.set(pasta)
        if not Path(pasta).is_dir():
            messagebox.showerror("Erro", f"Pasta não encontrada:\n{pasta}")
            return

        self._limpar_log()
        self.btn_rodar.config(state="disabled", text="Processando...")
        self._set_status("Processando...", WARNING)

        def run():
            try:
                pdf = Path(pdf_path)
                self._log(f"Arquivo: {pdf.name}")
                self._log(f"Destino: {pasta}\n")

                funcionarios = extrair_contracheques(pdf_path, self._log)

                if not funcionarios:
                    self._log("\nNenhum contracheque encontrado.")
                    self._set_status("Nenhum dado extraído — verifique o PDF.", ERROR)
                    messagebox.showwarning("Aviso", "Nenhum contracheque encontrado no PDF.")
                    return

                self._log(f"\n{len(funcionarios)} funcionário(s) extraído(s).\n")
                self._log("Validando dados extraídos...")
                alertas = validar_funcionarios(funcionarios, self._log)
                self._log("")

                saida = gerar_nome_saida(Path(pasta), pdf.stem)
                gerar_xlsx(funcionarios, str(saida), alertas=alertas)

                if alertas:
                    nomes_com_erro = "\n".join(
                        f"  [{a['cod']}] {a['nome']}: {'; '.join(a['erros'])}"
                        for a in alertas
                    )
                    messagebox.showwarning(
                        "Validação — inconsistências encontradas",
                        f"{len(alertas)} registro(s) com problema:\n\n{nomes_com_erro}\n\n"
                        "O arquivo foi gerado mesmo assim.\n"
                        "Confira o log para detalhes."
                    )

                self._log(f"Planilha salva: {saida.name}")
                self._set_status(f"Concluído — {saida.name}", SUCCESS)
                messagebox.showinfo(
                    "Concluído",
                    f"{len(funcionarios)} funcionário(s) extraído(s).\n\nArquivo:\n{saida}"
                )
            except Exception as e:
                self._log(f"\n[ERRO] {e}")
                self._set_status(f"Erro: {e}", ERROR)
                messagebox.showerror("Erro", str(e))
            finally:
                self.btn_rodar.config(state="normal", text="▶   EXTRAIR DADOS")

        threading.Thread(target=run, daemon=True).start()


if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()
